from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ---------- 样式 ----------
ARIAL = "Arial"
title_font = Font(name=ARIAL, bold=True, size=14, color="1F4E78")
hdr_font = Font(name=ARIAL, bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F4E78")
blue_font = Font(name=ARIAL, size=10, color="0000FF")          # 可改输入
black_font = Font(name=ARIAL, size=10, color="000000")
qual_fill = PatternFill("solid", fgColor="FFF2CC")            # 质变行 浅黄
normal_fill = PatternFill("solid", fgColor="FFFFFF")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="center")
ctr = Alignment(horizontal="center", vertical="center")

# ---------- Sheet1 技能配置总览 ----------
s1 = wb.active
s1.title = "Config"
s1["A1"] = "技能配置总览（蓝色=可改输入；改完告诉我，我按此优化 game.js）"
s1["A1"].font = title_font
headers = ["技能ID","名称","图标","元素","分类","最大等级","每级基础效果(汉语)",
           "质变节点(汉语)","基准值①(蓝)","基准值②(蓝)","每级系数/增量说明","代码绑定位置"]
s1.append([])
s1.append(headers)
hrow = 3
for c in range(1, len(headers)+1):
    cell = s1.cell(row=hrow, column=c)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap; cell.border = border

# skills 定义：base1 写入 I 列，供 Sheet2 公式引用
# qual: {等级: 汉语描述}
skills = [
 dict(id="damage", name="火力强化", icon="🔫", element="物理", category="bullet", maxLevel=99,
   perLevel="伤害 ×1.2（即 +20%）", qual={5:"穿透 +1"},
   base1=10, base2="", coef="每级 player.damage *= 1.2；Lv5 → bulletPiercing++",
   bind="game.js:824"),
 dict(id="fireRate", name="急速射击", icon="»", element="物理", category="bullet", maxLevel=99,
   perLevel="射击间隔 ×0.85（射速约 +17.6%）", qual={3:"射击间隔再 ×0.95", 5:"射击间隔再 ×0.95"},
   base1=500, base2="", coef="每级 fireRate *= 0.85；质变 Lv3/Lv5 各再 ×0.95",
   bind="game.js:826"),
 dict(id="bulletCount", name="多重射击", icon="🎯", element="物理", category="bullet", maxLevel=20,
   perLevel="子弹数 +1", qual={5:"命中后分裂迸射小弹"},
   base1=1, base2="", coef="每级 player.bulletCount++；Lv5 → _splitOnHit=true",
   bind="game.js:828"),
 dict(id="bulletSpeed", name="高速子弹", icon="💨", element="物理", category="bullet", maxLevel=99,
   perLevel="弹速 ×1.2（+20%）", qual={3:"弹速再 ×1.1", 5:"穿透 +1"},
   base1=10, base2="", coef="每级 bulletSpeed *= 1.2；质变 Lv3 ×1.1、Lv5 piercing++",
   bind="game.js:830"),
 dict(id="piercing", name="穿透弹", icon="🗡️", element="物理", category="bullet", maxLevel=30,
   perLevel="穿透 +1", qual={3:"穿透 +1", 5:"穿透命中溅射小范围"},
   base1=1, base2="", coef="每级 bulletPiercing++；Lv3 再 +1、Lv5 → _splash",
   bind="game.js:832"),
 dict(id="health", name="生命强化", icon="❤️", element="物理", category="buff", maxLevel=99,
   perLevel="最大生命 +20", qual={3:"每级额外生命 +10（Lv3 起每级共 +30）"},
   base1=100, base2="", coef="每级 maxHealth += 20；Lv3 起每级额外 +10",
   bind="game.js:834"),
 dict(id="explosive", name="爆炸弹", icon="💥", element="火", category="bullet", maxLevel=99,
   perLevel="爆炸半径 +20；AOE伤害 = 单发×(0.3+0.1×等级)", qual={3:"爆炸遗留火池(灼烧)", 5:"范围 +50% & 二次小爆"},
   base1=10, base2="", coef="半径=40+20×lv（Lv5×1.5）；AOE=伤害×(0.3+0.1×lv)（Lv5×1.5）",
   bind="game.js:836 / 3188-3207"),
 dict(id="lightning", name="闪电链", icon="⚡", element="雷", category="bullet", maxLevel=99,
   perLevel="闪电链数 +1", qual={3:"链目标 +2", 5:"链命中触发导电(链伤 +30%)"},
   base1=10, base2="", coef="链数=lv+1+（Lv3:+2）；链伤=伤害×0.4×(Lv5:×1.3)",
   bind="game.js:838 / 3209-3238"),
 dict(id="shield", name="护盾", icon="🛡️", element="物理", category="buff", maxLevel=8,
   perLevel="受到伤害 −10%", qual={3:"减伤额外 +5%", 5:"受击反弹 10%"},
   base1=0, base2="", coef="减伤=lv×0.1（封顶0.8）；Lv3 +0.05；Lv5 → _reflect",
   bind="game.js:840 / 958-962"),
 dict(id="crit", name="致命暴击", icon="💢", element="物理", category="buff", maxLevel=12,
   perLevel="暴击率 +5%", qual={3:"暴击伤害 +20%", 5:"暴击触发小爆炸"},
   base1=0, base2="", coef="暴击率=lv×0.05(封顶0.6)；暴伤倍率=1+lv×0.1(+Lv3 0.2)；Lv5 → _explode",
   bind="game.js:842 / 935-942"),
 dict(id="freeze", name="冰霜弹", icon="❄️", element="冰", category="bullet", maxLevel=8,
   perLevel="冰冻几率 +6%", qual={3:"冰冻残留减速", 5:"20% 几率眩晕"},
   base1=0, base2="", coef="冰冻率=lv×0.06(封顶0.5)；时长=1000+lv×120(封顶2600)；Lv5 → _stun",
   bind="game.js:844 / 943-948"),
 dict(id="slow", name="缓速弹", icon="🐌", element="冰", category="cc", maxLevel=8,
   perLevel="减速几率 +8%", qual={3:"减速更强(因子再 −0.1)", 5:"减速标记易感"},
   base1=0, base2="", coef="减速率=lv×0.08(封顶0.6)；因子=0.7−lv×0.04(Lv3 −0.1,下限0.3)；Lv5 → _conductive",
   bind="game.js:846 / 949-957"),
 dict(id="mine", name="地雷", icon="💣", element="物理", category="field", maxLevel=10,
   perLevel="地雷数量 +1；半径 +12", qual={3:"伤害 +50%", 5:"爆炸附加减速"},
   base1=10, base2="", coef="数量=2+lv；半径=40+12×lv；伤害=裸伤×(0.5+0.12×lv)×2(Lv3×1.5)",
   bind="game.js:849 / 3603-3634"),
 dict(id="oil", name="油渍", icon="🛢️", element="火", category="field", maxLevel=5,
   perLevel="油渍数量 +1；半径 ×1.1 复合成长", qual={3:"灼烧 DOT ×2", 5:"额外火池 +1 & 减速"},
   base1=10, base2="", coef="数量=1+lv(+Lv5 +1)；半径=52.2×1.1^(lv−1)；DOT=裸伤×(Lv3:×2)",
   bind="game.js:851 / 3636-3666"),
 dict(id="tornado", name="龙卷风", icon="🌪️", element="风", category="cc", maxLevel=10,
   perLevel="龙卷风半径 +20", qual={3:"牵引力 +50%", 5:"龙卷内风伤启用"},
   base1=10, base2="", coef="半径=140+20×lv；牵引=0.02×(1+0.3×lv)(Lv3×1.5)；Lv5 风DPS=裸伤×0.1/0.5s",
   bind="game.js:853 / 3668-3705"),
]

r = hrow + 1
for s in skills:
    s["s1row"] = r
    vals = [s["id"], s["name"], s["icon"], s["element"], s["category"], s["maxLevel"],
            s["perLevel"], "; ".join(f"Lv{k}: {v}" for k,v in s["qual"].items()),
            s["base1"], s["base2"], s["coef"], s["bind"]]
    for c, v in enumerate(vals, start=1):
        cell = s1.cell(row=r, column=c, value=v)
        cell.font = black_font; cell.alignment = wrap; cell.border = border
        if c in (9, 10):   # 基准值列 → 蓝色可改
            cell.font = blue_font
    r += 1

widths1 = [12, 12, 6, 8, 8, 8, 30, 34, 11, 11, 44, 22]
for i, w in enumerate(widths1, start=1):
    s1.column_dimensions[chr(64+i) if i<=26 else "A"].width = w
s1.freeze_panes = "A4"

# ---------- Sheet2 逐等级明细 ----------
s2 = wb.create_sheet("逐等级明细")
s2["A1"] = "逐等级数值明细（累计数值用 Excel 公式动态计算；黄底=质变节点；改 Sheet1 基准值后此处自动更新）"
s2["A1"].font = title_font
h2 = ["技能(图标)","等级","本等级效果(汉语)","主数值(累计)","副数值①(累计)","副数值②(累计)","质变"]
s2.append([])
s2.append(h2)
hrow2 = 3
for c in range(1, len(h2)+1):
    cell = s2.cell(row=hrow2, column=c)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = border

def f_damage(r, s):    return f"=Config!I{s['s1row']}*1.2^B{r}"
def f_fireRate(r, s):  return f"=1000/(Config!I{s['s1row']}*0.85^B{r}*IF(B{r}>=3,0.95,1)*IF(B{r}>=5,0.95,1))"
def f_bulletCount(r, s): return f"=Config!I{s['s1row']}+B{r}"
def f_bulletSpeed(r, s): return f"=Config!I{s['s1row']}*1.2^B{r}*IF(B{r}>=3,1.1,1)"
def f_piercing(r, s):   return f"=Config!I{s['s1row']}+B{r}+IF(B{r}>=3,1,0)"
def f_health(r, s):     return f"=Config!I{s['s1row']}+20*B{r}+IF(B{r}>=3,10*(B{r}-2),0)"
def f_explosive(r, s):  return f"=(40+20*B{r})*IF(B{r}>=5,1.5,1)"
def f_explosiveE(r, s): return f"=Config!I{s['s1row']}*(0.3+0.1*B{r})*IF(B{r}>=5,1.5,1)"
def f_lightning(r, s):  return f"=B{r}+1+IF(B{r}>=3,2,0)"
def f_lightningE(r, s): return f"=Config!I{s['s1row']}*0.4*IF(B{r}>=5,1.3,1)"
def f_shield(r, s):     return f"=B{r}*0.1+IF(B{r}>=3,0.05,0)"
def f_crit(r, s):       return f"=MIN(0.6,B{r}*0.05)"
def f_critE(r, s):      return f"=1+B{r}*0.1+IF(B{r}>=3,0.2,0)"
def f_freeze(r, s):     return f"=MIN(0.5,B{r}*0.06)"
def f_freezeE(r, s):    return f"=MIN(2600,1000+120*B{r})"
def f_slow(r, s):       return f"=MIN(0.6,B{r}*0.08)"
def f_slowE(r, s):      return f"=MAX(0.3,0.7-0.04*B{r}-IF(B{r}>=3,0.1,0))"
def f_mine(r, s):       return f"=2+B{r}"
def f_mineE(r, s):      return f"=40+12*B{r}"
def f_mineF(r, s):      return f"=Config!I{s['s1row']}*(0.5+0.12*B{r})*2*IF(B{r}>=3,1.5,1)"
def f_oil(r, s):        return f"=1+B{r}+IF(B{r}>=5,1,0)"
def f_oilE(r, s):       return f"=52.2*1.1^(B{r}-1)"
def f_oilF(r, s):       return f"=Config!I{s['s1row']}*(1+IF(B{r}>=3,1,0))"
def f_tornado(r, s):    return f"=140+20*B{r}"
def f_tornadoE(r, s):   return f"=0.02*(1+0.3*B{r})*IF(B{r}>=3,1.5,1)"
def f_tornadoF(r, s):   return f"=IF(B{r}>=5,Config!I{s['s1row']}*0.1,0)"

# 每个技能：主/副公式函数名映射
fmaps = {
 "damage":    (f_damage, None, None),
 "fireRate":  (f_fireRate, None, None),
 "bulletCount":(f_bulletCount, None, None),
 "bulletSpeed":(f_bulletSpeed, None, None),
 "piercing":  (f_piercing, None, None),
 "health":    (f_health, None, None),
 "explosive": (f_explosive, f_explosiveE, None),
 "lightning": (f_lightning, f_lightningE, None),
 "shield":    (f_shield, None, None),
 "crit":      (f_crit, f_critE, None),
 "freeze":    (f_freeze, f_freezeE, None),
 "slow":      (f_slow, f_slowE, None),
 "mine":      (f_mine, f_mineE, f_mineF),
 "oil":       (f_oil, f_oilE, f_oilF),
 "tornado":   (f_tornado, f_tornadoE, f_tornadoF),
}

row = hrow2 + 1
for s in skills:
    fd, fe, ff = fmaps[s["id"]]
    for n in range(1, s["maxLevel"]+1):
        is_qual = n in s["qual"]
        effect = s["qual"][n] if is_qual else s["perLevel"]
        a = s2.cell(row=row, column=1, value=f"{s['icon']}{s['name']}")
        b = s2.cell(row=row, column=2, value=n)
        c = s2.cell(row=row, column=3, value=effect)
        d = s2.cell(row=row, column=4, value=fd(row, s))
        e = s2.cell(row=row, column=5, value=fe(row, s) if fe else "-")
        f = s2.cell(row=row, column=6, value=ff(row, s) if ff else "-")
        g = s2.cell(row=row, column=7, value=("★Lv%d质变" % n) if is_qual else "")
        fill = qual_fill if is_qual else normal_fill
        for col in range(1, 8):
            cell = s2.cell(row=row, column=col)
            cell.border = border
            cell.fill = fill
            cell.font = black_font
            cell.alignment = wrap if col == 3 else ctr
        row += 1

widths2 = [14, 7, 40, 22, 22, 24, 12]
for i, w in enumerate(widths2, start=1):
    s2.column_dimensions[chr(64+i)].width = w
s2.freeze_panes = "A4"

# ---------- Sheet3 使用说明 ----------
s3 = wb.create_sheet("使用说明")
lines = [
 ("技能数值策划表 · 使用说明", title_font),
 ("", None),
 ("1. 本表由 game.js 的 SKILL_DEFS（15 个技能）如实导出，未做简化。", black_font),
 ("2. Sheet1「技能配置总览」：蓝色字体 = 你可以直接改的数值（基准值①/②、最大等级、每级系数说明）。", black_font),
 ("3. Sheet2「逐等级明细」：每一行=一个技能的一级；累计数值用 Excel 公式自动算出，改 Sheet1 基准会联动更新。", black_font),
 ("4. 黄底行 = 质变节点（Lv3 / Lv5），效果汉语已写明，与主效果不同。", black_font),
 ("5. 累计数值含义：", black_font),
 ("   - 乘法型（火力/弹速）：基准×每级系数^等级，例如火力 Lv3 = 10×1.2³ ≈ 17.28。", black_font),
 ("   - 加法型（生命/子弹/穿透）：基准 + 等级×增量。", black_font),
 ("   - 开关型（爆炸/闪电/护盾/暴击/冰/缓/地雷/油渍/龙卷）：无每级数值变化，数值在触发时按等级计算（已列公式）。", black_font),
 ("6. 你改完后把表发我（或直接告诉我改了哪几个数），我按此同步优化 game.js 的 SKILL_DEFS 与相关结算函数。", black_font),
 ("7. 注意：damage / fireRate / bulletSpeed / health / explosive / lightning 的 maxLevel=99 仅为上限，实战通常远低于此；", black_font),
 ("   真正影响手感的是 每级系数 与 质变阈值（Lv3/Lv5），这两处是配置重点。", black_font),
 ("8. 列说明：主数值=该技能核心数值；副数值①/②=多数值技能的次要数值（如半径/伤害/链数等）。", black_font),
]
rr = 1
for text, fnt in lines:
    cell = s3.cell(row=rr, column=1, value=text)
    if fnt: cell.font = fnt
    cell.alignment = wrap
    rr += 1
s3.column_dimensions["A"].width = 110

out = r"C:\Users\guoxiaoyu\WorkBuddy\game\zombie-hunter-game\技能数值表.xlsx"
wb.save(out)
print("SAVED", out, "rows2=", row-1)
